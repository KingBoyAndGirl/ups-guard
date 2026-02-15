"""历史记录 API"""
from fastapi import APIRouter, Query, HTTPException
from fastapi.responses import StreamingResponse
from typing import Optional
from datetime import datetime, timedelta
import csv
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from models import EventType
from services.history import get_history_service

router = APIRouter()


@router.get("/history/events")
async def get_events(
    days: int = Query(7, ge=1, le=365, description="查询最近几天的事件"),
    event_type: Optional[str] = Query(None, description="过滤事件类型")
):
    """获取历史事件"""
    history_service = await get_history_service()
    
    event_type_enum = None
    if event_type:
        try:
            event_type_enum = EventType(event_type)
        except ValueError:
            pass
    
    events = await history_service.get_events(days, event_type_enum)
    
    return {
        "events": [
            {
                "id": event.id,
                "event_type": event.event_type.value,
                "message": event.message,
                # Convert to UTC ISO format with Z suffix for proper timezone handling
                "timestamp": event.timestamp.isoformat().replace('+00:00', 'Z'),
                "metadata": event.metadata
            }
            for event in events
        ]
    }


@router.get("/history/metrics")
async def get_metrics(
    hours: int = Query(None, ge=1, le=720, description="查询最近几小时的指标"),
    minutes: int = Query(None, ge=1, le=60, description="查询最近几分钟的指标")
):
    """获取历史指标"""
    history_service = await get_history_service()
    
    # 优先使用 minutes 参数，否则使用 hours
    if minutes is not None:
        # 将分钟转换为小时（至少1小时以确保能获取数据）
        query_hours = max(1, minutes / 60)
        metrics = await history_service.get_metrics(query_hours)

        # 过滤到指定分钟范围内的数据
        if metrics:
            from datetime import timezone
            cutoff_time = datetime.now(timezone.utc) - timedelta(minutes=minutes)
            # 确保比较时都使用 UTC 时间
            metrics = [m for m in metrics if m.timestamp.replace(tzinfo=timezone.utc) >= cutoff_time]
    else:
        query_hours = hours if hours is not None else 24
        metrics = await history_service.get_metrics(query_hours)

    return {
        "metrics": [
            {
                "id": metric.id,
                # Convert to UTC ISO format with Z suffix for proper timezone handling
                "timestamp": metric.timestamp.isoformat().replace('+00:00', 'Z'),
                "battery_charge": metric.battery_charge,
                "battery_runtime": metric.battery_runtime,
                "input_voltage": metric.input_voltage,
                "output_voltage": metric.output_voltage,
                "load_percent": metric.load_percent,
                "temperature": metric.temperature
            }
            for metric in metrics
        ]
    }


@router.get("/history/export")
async def export_history(
    format: str = Query("csv", description="导出格式: csv 或 xlsx"),
    type: str = Query("all", description="数据类型: events, metrics, 或 all"),
    start_date: Optional[str] = Query(None, description="开始日期 (ISO格式)"),
    end_date: Optional[str] = Query(None, description="结束日期 (ISO格式)")
):
    """导出历史数据"""
    # 验证格式
    if format not in ["csv", "xlsx"]:
        raise HTTPException(status_code=400, detail="format 必须是 'csv' 或 'xlsx'")
    
    if type not in ["events", "metrics", "all"]:
        raise HTTPException(status_code=400, detail="type 必须是 'events', 'metrics' 或 'all'")
    
    # 解析日期范围（统一使用 naive datetime 进行比较）
    try:
        if start_date:
            start_dt = datetime.fromisoformat(start_date.replace('Z', '+00:00'))
            # 转换为 naive datetime 以便与数据库中的时间比较
            if start_dt.tzinfo is not None:
                start_dt = start_dt.replace(tzinfo=None)
        else:
            start_dt = datetime.now() - timedelta(days=30)  # 默认最近30天
        
        if end_date:
            end_dt = datetime.fromisoformat(end_date.replace('Z', '+00:00'))
            # 转换为 naive datetime 以便与数据库中的时间比较
            if end_dt.tzinfo is not None:
                end_dt = end_dt.replace(tzinfo=None)
        else:
            end_dt = datetime.now()
        
        # 限制最多90天
        if (end_dt - start_dt).days > 90:
            raise HTTPException(status_code=400, detail="导出时间范围不能超过90天")
    
    except ValueError as e:
        raise HTTPException(status_code=400, detail=f"日期格式错误: {str(e)}")
    
    history_service = await get_history_service()
    
    # 生成文件名
    date_str = datetime.now().strftime('%Y%m%d-%H%M%S')
    filename = f"ups-guard-{type}-{date_str}.{format}"
    
    # 根据格式导出
    if format == "csv":
        return await export_csv(history_service, type, start_dt, end_dt, filename)
    else:
        return await export_xlsx(history_service, type, start_dt, end_dt, filename)


async def export_csv(history_service, data_type: str, start_dt: datetime, end_dt: datetime, filename: str):
    """导出为 CSV 格式"""
    output = io.StringIO()
    
    if data_type in ["events", "all"]:
        # 导出事件
        days = (end_dt - start_dt).days + 1
        events = await history_service.get_events(days, None)
        
        # 过滤日期范围
        filtered_events = [e for e in events if start_dt <= e.timestamp <= end_dt]
        
        writer = csv.writer(output)
        writer.writerow(['时间', '事件类型', '描述'])
        
        for event in filtered_events:
            writer.writerow([
                event.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                event.event_type.value,
                event.message
            ])
        
        if data_type == "all":
            writer.writerow([])  # 空行分隔
    
    if data_type in ["metrics", "all"]:
        # 导出指标
        hours = int((end_dt - start_dt).total_seconds() / 3600) + 1
        metrics = await history_service.get_metrics(hours)
        
        # 过滤日期范围
        filtered_metrics = [m for m in metrics if start_dt <= m.timestamp <= end_dt]
        
        writer = csv.writer(output)
        if data_type == "all":
            writer.writerow(['指标数据'])
        writer.writerow(['时间', '电池电量(%)', '输入电压(V)', '输出电压(V)', '负载(%)', '温度(°C)', '运行时间(秒)'])
        
        for metric in filtered_metrics:
            writer.writerow([
                metric.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                metric.battery_charge if metric.battery_charge is not None else '',
                metric.input_voltage if metric.input_voltage is not None else '',
                metric.output_voltage if metric.output_voltage is not None else '',
                metric.load_percent if metric.load_percent is not None else '',
                metric.temperature if metric.temperature is not None else '',
                metric.battery_runtime if metric.battery_runtime is not None else ''
            ])
    
    # 创建响应
    output.seek(0)
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


async def export_xlsx(history_service, data_type: str, start_dt: datetime, end_dt: datetime, filename: str):
    """导出为 Excel 格式"""
    wb = Workbook()
    
    # 样式定义
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    
    if data_type in ["events", "all"]:
        # 创建事件工作表
        ws_events = wb.active
        ws_events.title = "事件记录"
        
        # 设置表头
        headers = ['时间', '事件类型', '描述']
        ws_events.append(headers)
        
        # 应用表头样式
        for cell in ws_events[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # 获取并写入数据
        days = (end_dt - start_dt).days + 1
        events = await history_service.get_events(days, None)
        filtered_events = [e for e in events if start_dt <= e.timestamp <= end_dt]
        
        for event in filtered_events:
            ws_events.append([
                event.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                event.event_type.value,
                event.message
            ])
        
        # 自动调整列宽
        for column in ws_events.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_events.column_dimensions[column_letter].width = adjusted_width
    
    if data_type in ["metrics", "all"]:
        # 创建指标工作表
        if data_type == "all":
            ws_metrics = wb.create_sheet(title="指标数据")
        else:
            ws_metrics = wb.active
            ws_metrics.title = "指标数据"
        
        # 设置表头
        headers = ['时间', '电池电量(%)', '输入电压(V)', '输出电压(V)', '负载(%)', '温度(°C)', '运行时间(秒)']
        ws_metrics.append(headers)
        
        # 应用表头样式
        for cell in ws_metrics[1]:
            cell.font = header_font
            cell.fill = header_fill
        
        # 获取并写入数据
        hours = int((end_dt - start_dt).total_seconds() / 3600) + 1
        metrics = await history_service.get_metrics(hours)
        filtered_metrics = [m for m in metrics if start_dt <= m.timestamp <= end_dt]
        
        for metric in filtered_metrics:
            ws_metrics.append([
                metric.timestamp.strftime('%Y-%m-%d %H:%M:%S'),
                metric.battery_charge,
                metric.input_voltage,
                metric.output_voltage,
                metric.load_percent,
                metric.temperature,
                metric.battery_runtime
            ])
        
        # 自动调整列宽
        for column in ws_metrics.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_metrics.column_dimensions[column_letter].width = adjusted_width
    
    # 保存到内存
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        iter([output.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/history/events")
async def create_event(event_data: dict):
    """记录事件（支持前端事件记录）"""
    from pydantic import BaseModel
    
    # 验证必需字段
    if 'event_type' not in event_data or 'message' not in event_data:
        raise HTTPException(status_code=400, detail="Missing required fields: event_type and message")
    
    try:
        # 验证事件类型
        event_type = EventType(event_data['event_type'])
    except ValueError:
        raise HTTPException(status_code=400, detail=f"Invalid event_type: {event_data['event_type']}")
    
    # 记录事件
    history_service = await get_history_service()
    metadata = event_data.get('metadata', {})
    
    try:
        await history_service.add_event(
            event_type=event_type,
            message=event_data['message'],
            metadata=metadata
        )
        return {"status": "success", "message": "Event recorded"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Failed to record event: {str(e)}")
